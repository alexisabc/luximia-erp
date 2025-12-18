
import os
import datetime
from decimal import Decimal
from django.db import transaction
from django.utils.text import slugify
from .models import (
    Nomina, ReciboNomina, Empleado, ConceptoNomina, DetalleReciboItem, 
    TipoConcepto, RazonSocial, NominaCentralizada
)
import openpyxl

try:
    from fuzzywuzzy import process
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False

class NominaImporter:
    def __init__(self, stdout=None):
        self.stdout = stdout

    def log(self, message, style=None):
        if self.stdout:
            if style:
                self.stdout.write(style(message))
            else:
                self.stdout.write(message)
        else:
            print(message) 

    def normalize_name(self, name):
        if not name: return ""
        return str(name).upper().strip()

    def get_column_mapping(self):
        """
        Define el mapeo entre campos del modelo NominaCentralizada y posibles headers en Excel.
        """
        return {
            'codigo': ['CODIGO', 'CÓDIGO', 'ID', 'NO'],
            'nombre': ['NOMBRE', 'EMPLEADO', 'TRABAJADOR', 'NOMBRE COMPLETO'],
            'departamento': ['DEPTO', 'DEPARTAMENTO'],
            'puesto': ['PUESTO', 'CARGO'],
            'neto_mensual': ['NETO MENSUAL'],
            'sueldo_diario': ['SDO', 'SALARIO DIARIO', 'SD', 'CUOTA DIARIA'],
            'dias_trabajados': ['DIAS', 'DÍAS', 'DIAS TRAB'],
            'sueldo': ['SUELDO', 'SALARIO'], # Sueldo del periodo
            'vacaciones': ['VACACIONES', 'VAC'],
            'prima_vacacional': ['PRIMA VACACIONAL', 'PV'],
            'aguinaldo': ['AGUINALDO', 'AGUIN'],
            'retroactivo': ['RETROACTIVO', 'DIA TRABAJADO PEND', 'DIAS PENDIENTES'],
            'subsidio': ['SUBSIDIO', 'SUB EMPLEO'],
            'total_percepciones': ['TOTAL PERCEPCIONES', 'SUMA PERCEPCIONES', 'TOTAL PERC', 'TOT PERCEP', 'PERCEPCIONES', 'T. PERCEPCIONES', 'T PERCEPCIONES', 'TOTAL DE PERCEPCIONES'],
            'isr': ['ISR', 'IMPUESTO'],
            'imss': ['IMSS', 'SEGURO SOCIAL'],
            'prestamo': ['PRESTAMO', 'PRÉSTAMO'],
            'infonavit': ['INFONAVIT', 'CREDITO VIVIENDA'],
            'total_deducciones': ['TOTAL DEDUCCIONES', 'SUMA DEDUCCIONES', 'T. DEDUCCIONES', 'DEDUCCIONES', 'REDUCCIONES', 'TOTAL DE DEDUCCIONES', 'DEDUCCIONES TOTAL'],
            'neto': ['NETO', 'A PAGAR', 'NETO A PAGAR', 'LIQUIDO'],
            'isn': ['ISN', '0.04', '3%', 'IMPUESTO SOBRE NOMINA'],
            'previo_costo_social': ['PREVIO COSTO SOCIAL', 'COSTO SOCIAL'],
            'total_carga_social': ['TOTAL CARGA SOCIAL', 'CARGA SOCIAL'],
            'total_nomina': ['TOTAL NOMINA', 'GRAN TOTAL'],
            'nominas_y_costos': ['NOMINAS Y COSTOS TRIBUTARIO', 'NOMINA + COSTO'],
            'comision': ['COMISION', 'COMISIÓN'],
            'sub_total': ['SUB-TOTAL', 'SUBTOTAL'],
            'iva': ['IVA'],
            'total_facturacion': ['TOTAL FACTURACION', 'FACTURACION', 'TOTAL FACTURA']
        }

    def find_header_row(self, ws):
        # ... (Mantener lógica existente)
        employee_keywords = ["CODIGO", "CÓDIGO", "NOMBRE", "EMPLEADO", "TRABAJADOR"]
        best_row = None
        best_score = 0

        for row_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True), start=1):
            row_str = [str(c).upper().strip() for c in row if c]
            matches = sum(1 for k in employee_keywords if any(k == cell_val or k in cell_val for cell_val in row_str))
            
            score = matches
            if "CÓDIGO" in row_str or "CODIGO" in row_str:
                score += 5 
            if "NOMBRE" in row_str:
                score += 5

            if score > best_score and score >= 2: 
               best_score = score
               best_row = (row_idx, [str(c).strip() if c else "" for c in row])
        
        if best_row:
             return best_row
        return None, None

    def process_sheet_centralized(self, ws, sheet_name, anio, dry_run):
        header_row_idx, headers = self.find_header_row(ws)
        if not header_row_idx:
            return {'sheet': sheet_name, 'status': 'skipped', 'reason': 'No headers found'}

        # Mapear columnas a indices
        field_indices = {}
        mapping = self.get_column_mapping()
        
        # DEBUG: Imprimir headers encontrados para diagnostico
        print(f"DEBUG Headers detectados hoja {sheet_name}: {headers}")
        
        # Exclusiones explícitas para evitar falsos positivos
        exclusions = {
            'neto': ['MENSUAL', 'ANUAL'],
            'total_percepciones': ['GRAV', 'EXEN', 'DEDUCCIONES', 'NETO', 'TOTAL DEDUCCIONES'],
            'total_deducciones': ['PERCEPCIONES']
        }

        # -------------------------------------------------------------------------
        # Header Enrichment Logic (Manejo de Celdas Combinadas Vertical/Horizontal)
        # -------------------------------------------------------------------------
        
        # Estrategia: Buscar "Super Headers" en la fila ANTERIOR a la detectada (ej. "PERCEPCIONES" arriba de "TOTAL")
        headers_super = []
        if header_row_idx > 1:
            try:
                # openpyxl iter_rows indices are 1-based.
                # headers detected at header_row_idx.
                # We want header_row_idx - 1.
                row_gen = ws.iter_rows(min_row=header_row_idx - 1, max_row=header_row_idx - 1, values_only=True)
                headers_super = next(row_gen, [])
            except:
                headers_super = []

        rich_headers = []
        last_super = ""
        
        # Normalizar longitud
        max_len = max(len(headers), len(headers_super))
        
        for i in range(max_len):
            h_main = str(headers[i]).strip() if i < len(headers) and headers[i] else ""
            h_super = str(headers_super[i]).strip() if i < len(headers_super) and headers_super[i] else ""
            
            # Forward fill para h_super (el encabezado agrupador)
            if h_super:
                last_super = h_super
            elif not h_super and last_super:
                # Si estamos en un hueco, usamos el ultimo super header PERO solo si h_main tiene valor
                # (es decir, es una columna real bajo el grupo) o si parece continuidad.
                # En Excel merged cells: [PERCEPCIONES][ ][ ]
                #                        [Sueldo      ][Bono][Total]
                pass 
            
            # Reset last_super si encontramos un bloque nuevo y vacío o separador grande? 
            # Simplificación: El fill se mantiene hasta que cambie o se acabe la tabla visualmente.
            # Riesgo: que el fill se extienda a DEDUCCIONES si DEDUCCIONES no tiene titulo arriba.
            # Normalmente los super headers son contiguos: [PERCEPCIONES...][DEDUCCIONES...]
            # Si h_super cambia (incluso a vacio si el grupo terminó) deberiamos limpiar. 
            # Dificil saber si "vacío" es termino de grupo o merge. 
            # Heuristica: Si h_super esta vacio, usamos last_super.
            
            current_super = h_super if h_super else last_super
            
            # Construir header rico: "PERCEPCIONES TOTAL"
            if current_super and h_main: 
                full_h = f"{current_super} {h_main}".strip()
            else:
                full_h = h_main if h_main else current_super
                
            rich_headers.append(full_h)

        print(f"DEBUG Rich Headers (UPWARD scanning) hoja {sheet_name}: {rich_headers}")

        # Mapear columnas a indices USANDO rich_headers
        field_indices = {}
        mapping = self.get_column_mapping()
        
        # Exclusiones explícitas para evitar falsos positivos
        exclusions = {
            'neto': ['MENSUAL', 'ANUAL'],
            'total_percepciones': ['GRAV', 'EXEN', 'DEDUCCIONES', 'NETO', 'TOTAL DEDUCCIONES'],
            'total_deducciones': ['NETO', 'LIQUIDO', 'A PAGAR']
        }

        # Usamos rich_headers para el mapeo
        # ESTRATEGIA: Mapear primero campos específicos, luego genéricos
        priority_fields = ['total_percepciones', 'total_deducciones', 'neto', 'neto_mensual']
        
        # Primera pasada: campos prioritarios con coincidencia exacta
        for field in priority_fields:
            if field in mapping and field not in field_indices:
                variants = mapping[field]
                for idx, header in enumerate(rich_headers):
                    h_upper = str(header).upper().strip().replace('\n', ' ').replace('  ', ' ')
                    
                    # Verificar exclusiones
                    if field in exclusions:
                        if any(excl in h_upper for excl in exclusions[field]):
                            continue
                    
                    # Solo coincidencia exacta en primera pasada
                    if h_upper in variants:
                        field_indices[field] = idx
                        print(f"DEBUG MAPEO EXACTO: {field} -> '{header}' (col {idx})")
                        break
        
        # Segunda pasada: todos los campos con coincidencia parcial
        for idx, header in enumerate(rich_headers):
            h_upper = str(header).upper().strip().replace('\n', ' ').replace('  ', ' ')
            
            for field, variants in mapping.items():
                if field not in field_indices: 
                    # Verificar exclusiones
                    if field in exclusions:
                        if any(excl in h_upper for excl in exclusions[field]):
                            continue

                    if h_upper in variants: # Coincidencia exacta
                        field_indices[field] = idx
                        print(f"DEBUG MAPEO: {field} -> '{header}' (col {idx})")
                    elif any(v in h_upper for v in variants): # Coincidencia parcial
                        field_indices[field] = idx
                        print(f"DEBUG MAPEO PARCIAL: {field} -> '{header}' (col {idx})")


        # Fallback para ISN si es '0.04' literal
        if 'isn' not in field_indices:
             for idx, header in enumerate(rich_headers):
                 if '0.04' in header or '3%' in header:
                     field_indices['isn'] = idx
                     break
        
        print(f"DEBUG Indices detectados hoja {sheet_name}: {field_indices}")

        # Validar colunmna minima requerida
        if 'nombre' not in field_indices:
            return {'sheet': sheet_name, 'status': 'skipped', 'reason': 'Columna NOMBRE no encontrada'}

        processed_count = 0
        created_records = []
        errors = []

        # Iniciamos lectura. 
        # NOTA: headersdetected index es la ultima fila de headers. +1 es datos.
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Obtener nombre
            col_nombre = field_indices['nombre']
            nombre_val = row[col_nombre] if col_nombre < len(row) else None
            
            if not nombre_val: continue
            
            nombre_upper = str(nombre_val).upper()
            if any(x in nombre_upper for x in ["NUEVOS", "BAJAS", "FINIQUITO", "TOTAL", "SUMA"]):
                 if any(x in nombre_upper for x in ["NUEVOS", "BAJAS", "FINIQUITO"]):
                     break # Stop reading
                 continue # Skip summary rows

            # Lógica especial para TOTAL PERCEPCIONES con celdas combinadas
            col_percep = field_indices.get('total_percepciones')
            val_percep = None
            if col_percep is not None:
                # Determinar rango de fusión (headers vacíos a la derecha)
                scan_range = 1
                for offset in range(1, 5): # Check next 4 cols
                    check_idx = col_percep + offset
                    if check_idx < len(headers):
                         neighbor_header = str(headers[check_idx]).strip()
                         if neighbor_header == 'None' or neighbor_header == '':
                             scan_range += 1
                         else:
                             break
                    else:
                        break
                
                # Buscar valores en el rango
                candidates = []
                for offset in range(scan_range):
                     p_idx = col_percep + offset
                     if p_idx < len(row):
                         p_val = row[p_idx]
                         # Try cleaning
                         if isinstance(p_val, (int, float, Decimal)):
                             candidates.append(p_val)
                         elif isinstance(p_val, str):
                             try:
                                 v_clean = p_val.replace('$', '').replace(',', '').strip()
                                 if v_clean and v_clean != '-':
                                     candidates.append(Decimal(v_clean))
                             except: pass
                
                if candidates:
                    # Asumimos que el Total es el valor máximo en el rango (o el único)
                    val_percep = max(candidates)

            # Lógica especial para TOTAL DEDUCCIONES con celdas combinadas (similar a percepciones)
            col_deduc = field_indices.get('total_deducciones')
            val_deduc = None
            occupied_indices = set(field_indices.values())

            if col_deduc is not None:
                scan_range = 1
                for offset in range(1, 4): # Check next 3 cols
                    check_idx = col_deduc + offset
                    
                    # Prevent bleed into mapped fields (e.g., NETO)
                    if check_idx in occupied_indices:
                        break

                    if check_idx < len(headers):
                         # OJO: headers variable contiene headers simples. 
                         # Si rich_headers se usó para el index, el index apunta a la col fisica.
                         # Verificamos si headers[check_idx] es vacío (merge).
                         neighbor_header = str(headers[check_idx]).strip()
                         if neighbor_header == 'None' or neighbor_header == '':
                             scan_range += 1
                         else:
                             break
                    else:
                        break
                
                candidates = []
                for offset in range(scan_range):
                     p_idx = col_deduc + offset
                     if p_idx < len(row):
                         p_val = row[p_idx]
                         if isinstance(p_val, (int, float, Decimal)):
                             candidates.append(p_val)
                         elif isinstance(p_val, str):
                             try:
                                 v_clean = p_val.replace('$', '').replace(',', '').strip()
                                 if v_clean and v_clean != '-':
                                     candidates.append(Decimal(v_clean))
                             except: pass
                
                if candidates:
                    val_deduc = max(candidates)

            # Extraer datos
            data = {
                'esquema': 'FISCAL',
                'tipo': 'QUINCENAL',
                'periodo': '1', # Hardcoded por ahora según request
                'empresa': sheet_name,
                'nombre': nombre_val,
                # Metadata de auditoría
                'archivo_origen': 'Carga Masiva' 
            }

            for field, col_idx in field_indices.items():
                if field == 'nombre': continue
                
                # Usar valor pre-calculado para Total Percepciones
                if field == 'total_percepciones' and val_percep is not None:
                     data[field] = val_percep
                     continue

                # Usar valor pre-calculado para Total Deducciones
                if field == 'total_deducciones' and val_deduc is not None:
                     data[field] = val_deduc
                     continue

                val = row[col_idx] if col_idx < len(row) else None
                
                # Limpiar y convertir valor
                if val is None:
                    clean_val = 0
                elif isinstance(val, (int, float, Decimal)):
                    clean_val = val
                elif isinstance(val, str):
                    v_str = val.replace('$', '').replace(',', '').strip()
                    if not v_str:
                        clean_val = 0
                    elif v_str == '-':
                        clean_val = 0
                    else:
                        try:
                            clean_val = Decimal(v_str)
                        except:
                            clean_val = v_str # Dejar como string si falla (para campos no numéricos)
                else:
                    clean_val = val

                # Validar campos numéricos obligatorios a 0 si falló conversión string
                if field not in ['codigo', 'departamento', 'puesto', 'empresa', 'nombre', 'esquema', 'tipo', 'periodo']:
                    if not isinstance(clean_val, (int, float, Decimal)):
                        clean_val = 0
                
                data[field] = clean_val

            # Crear objeto
            try:
                # Usamos update_or_create para evitar duplicados si se corre varias veces sobre el mismo periodo/empresa/empleado
                # aunque el usuario pidió "funnel", update_or_create es más seguro para re-importaciones.
                # Clave única lógica: Empresa + Periodo + Nombre (o codigo si existe)
                
                defaults = data.copy()
                # Remove keys used for lookup
                lookup_keys = ['empresa', 'periodo', 'nombre']
                for k in lookup_keys:
                    defaults.pop(k, None)

                obj, created = NominaCentralizada.objects.update_or_create(
                    empresa=data['empresa'],
                    periodo=data['periodo'],
                    nombre=data['nombre'],
                    defaults=defaults
                )
                
                processed_count += 1
                created_records.append(f"{data['nombre']} ({data['empresa']})")
                
            except Exception as e:
                errors.append(f"Error guardando {data['nombre']}: {str(e)}")

        return {
            'sheet': sheet_name,
            'status': 'success',
            'processed': processed_count,
            'processed_names': created_records, # Reutilizamos este campo para el UI
            'errors': errors
        }

    def process_file(self, file_path_or_obj, anio=2025, dry_run=False):
        try:
            wb = openpyxl.load_workbook(file_path_or_obj, data_only=True)
        except Exception as e:
            self.log(f"Error al abrir Excel: {e}")
            raise e

        # Ya no necesitamos empleados_db para este modo "single table dump" si no hacemos matching
        # El usuario pidió "guardar la información de manera centralizada en 1 sola tabla"
        # No mencionó vincular con Empleado model existente, solo "historico"
        
        results = []

        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                normalized_name = sheet_name.upper().strip()
                
                # Filtros básicos
                target_sheets = ["LUXIMIA", "SHARK", "PETREOS"]
                if any(exclude in normalized_name for exclude in ["RESUMEN", "YUKAWA"]):
                     continue
                if not any(target in normalized_name for target in target_sheets):
                     continue

                self.log(f"Procesando pestaña (Centralizada): {sheet_name}")
                ws = wb[sheet_name]
                
                # USAR NUEVO MÉTODO
                summary = self.process_sheet_centralized(ws, sheet_name, anio, dry_run)
                results.append(summary)

            if dry_run:
                transaction.set_rollback(True)
        
        return results
