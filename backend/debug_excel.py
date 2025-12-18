
import openpyxl
import sys

file_path = "../.agent/workflows/02.FISCAL QNA DEL 16-31 ENERO 2025.xlsx"

try:
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet in ["LUXIMIA", "SHARK", "PETREOS"]:
        if sheet in wb.sheetnames:
            print(f"\n--- SHEET: {sheet} ---")
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
                # Filter out None values for cleaner output
                cleaned_row = [str(c).strip() for c in row if c is not None]
                print(f"Row {i}: {cleaned_row}")
        else:
            print(f"Sheet {sheet} not found")

except Exception as e:
    print(f"Error: {e}")
