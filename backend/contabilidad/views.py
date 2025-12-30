from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, DecimalField
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay, Coalesce
from django.utils import timezone
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils.dateparse import parse_date
from .models import (
    CertificadoDigital,
    CuentaContable,
    BuzonMensaje,
    OpinionCumplimiento,
    # ... existing imports implicitly handled by viewsets logic below if needed, 
    # but strictly speaking I should check if I need to add specific imports if I replace the whole file or parts.
    # The user asked to "Help create logic...". I will append/insert the new ViewSets.
)
from .serializers import (
    CertificadoDigitalSerializer,
    BuzonMensajeSerializer,
    OpinionCumplimientoSerializer,
)
from .services.reportes import ReporteFinancieroService
from .services.sat_integration import SATIntegrationService
from django.http import HttpResponse
import openpyxl




from core.permissions import HasPermissionForAction
from .utils import sincronizar_tipo_cambio_banxico
from .models import (
    Banco,
    Proyecto,
    UPE,
    Cliente,
    Pago,
    Moneda,
    MetodoPago,
    Presupuesto,
    Contrato,
    TipoCambio,
    Vendedor,
    FormaPago,
    PlanPago,
    PlanPago,
    EsquemaComision,
    CuentaContable,
    CentroCostos,
    Poliza,
    DetallePoliza,
    Factura,
)

from .serializers import (
    BancoSerializer,
    ProyectoSerializer,
    UPESerializer,
    ClienteSerializer,
    PagoSerializer,
    MonedaSerializer,
    MetodoPagoSerializer,
    PresupuestoSerializer,
    ContratoSerializer,
    TipoCambioSerializer,
    VendedorSerializer,
    FormaPagoSerializer,
    PlanPagoSerializer,
    EsquemaComisionSerializer,
    CuentaContableSerializer,
    CentroCostosSerializer,
    PolizaSerializer,
    DetallePolizaSerializer,
    FacturaSerializer,
)


MAX_QUERY_TOKENS = 200
MAX_RESPONSE_TOKENS = 300
MAX_CONTEXT_TOKENS = 1000


def _truncate_words(text: str, limit: int) -> str:
    tokens = text.split()
    if len(tokens) <= limit:
        return text
    return " ".join(tokens[:limit])


from core.views import BaseViewSet

class ContabilidadBaseViewSet(BaseViewSet):
    permission_classes = [HasPermissionForAction]


class BancoViewSet(ContabilidadBaseViewSet):
    queryset = Banco.objects.all().order_by("id")
    serializer_class = BancoSerializer


class ProyectoViewSet(ContabilidadBaseViewSet):
    queryset = Proyecto.objects.all().order_by("id")
    serializer_class = ProyectoSerializer


class UPEViewSet(ContabilidadBaseViewSet):
    queryset = UPE.objects.select_related("proyecto", "moneda").all().order_by("id")
    serializer_class = UPESerializer


class ClienteViewSet(ContabilidadBaseViewSet):
    queryset = Cliente.objects.all().order_by("id")
    serializer_class = ClienteSerializer


class PagoViewSet(ContabilidadBaseViewSet):
    queryset = Pago.objects.all().order_by("id")
    serializer_class = PagoSerializer


class MonedaViewSet(ContabilidadBaseViewSet):
    queryset = Moneda.objects.all().order_by("id")
    serializer_class = MonedaSerializer


class MetodoPagoViewSet(ContabilidadBaseViewSet):
    queryset = MetodoPago.objects.all().order_by("id")
    serializer_class = MetodoPagoSerializer


class TipoCambioViewSet(ContabilidadBaseViewSet):
    queryset = TipoCambio.objects.all().order_by("id")
    serializer_class = TipoCambioSerializer


class TipoDeCambioSATViewSet(ContabilidadBaseViewSet):
    """
    Vista específica para tipos de cambio de BANXICO (SAT).
    Endpoint: /contabilidad/tipos-de-cambio/
    """
    queryset = TipoCambio.objects.filter(escenario='BANXICO').order_by("-fecha")
    serializer_class = TipoCambioSerializer

    @action(detail=False, methods=['post'])
    def actualizar(self, request):
        from datetime import date
        resultado = sincronizar_tipo_cambio_banxico(date.today())
        
        if resultado['success']:
             return Response({"mensaje": resultado['message']})
        return Response({"error": resultado['message']}, status=status.HTTP_400_BAD_REQUEST)


class VendedorViewSet(ContabilidadBaseViewSet):
    queryset = Vendedor.objects.all().order_by("id")
    serializer_class = VendedorSerializer


class FormaPagoViewSet(ContabilidadBaseViewSet):
    queryset = FormaPago.objects.all().order_by("id")
    serializer_class = FormaPagoSerializer


class PlanPagoViewSet(ContabilidadBaseViewSet):
    queryset = PlanPago.objects.all().order_by("id")
    serializer_class = PlanPagoSerializer


class EsquemaComisionViewSet(ContabilidadBaseViewSet):
    queryset = EsquemaComision.objects.all().order_by("id")
    serializer_class = EsquemaComisionSerializer


class PresupuestoViewSet(ContabilidadBaseViewSet):
    queryset = Presupuesto.objects.all().order_by("id")
    serializer_class = PresupuestoSerializer


class ContratoViewSet(ContabilidadBaseViewSet):
    queryset = Contrato.objects.all().order_by("id")
    serializer_class = ContratoSerializer





class CuentaContableViewSet(ContabilidadBaseViewSet):
    queryset = CuentaContable.objects.all().order_by("codigo")
    serializer_class = CuentaContableSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        if show_inactive and hasattr(CuentaContable, 'all_objects'):
            queryset = CuentaContable.all_objects.all().order_by("codigo")
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(codigo__icontains=search) | 
                models.Q(nombre__icontains=search)
            )
        return queryset

    @action(detail=False, methods=['get'])
    def exportar(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="cuentas_contables.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas"
        ws.append(['Código', 'Nombre', 'Tipo', 'Naturaleza', 'Nivel', 'Agrupador SAT'])
        
        for p in self.filter_queryset(self.get_queryset()):
            ws.append([p.codigo, p.nombre, p.tipo, p.naturaleza, p.nivel, p.codigo_agrupador])
            
        wb.save(response)
        return response

class CentroCostosViewSet(ContabilidadBaseViewSet):
    queryset = CentroCostos.objects.all().order_by("codigo")
    serializer_class = CentroCostosSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        if show_inactive and hasattr(CentroCostos, 'all_objects'):
            queryset = CentroCostos.all_objects.all().order_by("codigo")

        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(codigo__icontains=search) | 
                models.Q(nombre__icontains=search)
            )
        return queryset

    @action(detail=False, methods=['get'])
    def exportar(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="centros_costos.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Centros de Costos"
        ws.append(['Código', 'Nombre'])
        
        for p in self.filter_queryset(self.get_queryset()):
            ws.append([p.codigo, p.nombre])
            
        wb.save(response)
        return response

class PolizaViewSet(ContabilidadBaseViewSet):
    queryset = Poliza.objects.all().order_by("-fecha", "-numero")
    serializer_class = PolizaSerializer

class DetallePolizaViewSet(ContabilidadBaseViewSet):
    queryset = DetallePoliza.objects.all()
    serializer_class = DetallePolizaSerializer



class FacturaViewSet(ContabilidadBaseViewSet):
    queryset = Factura.objects.all().order_by("-fecha_emision")
    serializer_class = FacturaSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        
        if show_inactive and hasattr(Factura, 'all_objects'):
            queryset = Factura.all_objects.all().order_by("-fecha_emision")
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(uuid__icontains=search) | 
                models.Q(serie__icontains=search) |
                models.Q(folio__icontains=search) |
                models.Q(receptor_nombre__icontains=search) |
                models.Q(receptor_rfc__icontains=search)
            )
            
        return queryset

    @action(detail=False, methods=['get'])
    def exportar(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="facturas.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        
        headers = ['UUID', 'Serie', 'Folio', 'Fecha', 'Receptor', 'RFC Receptor', 'Total', 'Estado', 'Moneda']
        ws.append(headers)
        
        facturas = self.filter_queryset(self.get_queryset())
        
        for f in facturas:
            ws.append([
                f.uuid, f.serie, f.folio, f.fecha_emision.strftime('%Y-%m-%d') if f.fecha_emision else '',
                f.receptor_nombre, f.receptor_rfc, f.total, f.estado_sat, 
                f.moneda.codigo if f.moneda else ''
            ])
            
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='upload-xml')
    def upload_xml(self, request):
        """
        Sube y procesa uno o múltiples archivos XML (CFDI).
        """
        from .services.factura_service import FacturaService
        
        archivos = request.FILES.getlist('xmls')
        if not archivos:
            return Response({"detalle": "No se enviaron archivos."}, status=status.HTTP_400_BAD_REQUEST)

        resultados = {
            "procesados": 0,
            "errores": 0,
            "duplicados": 0,
            "detalles": []
        }

        for archivo in archivos:
            nombre_archivo = archivo.name
            
            # Procesar usando el servicio
            resultado = FacturaService.procesar_factura(archivo)
            
            if resultado['status'] == 'success':
                resultados['procesados'] += 1
                resultados['detalles'].append({
                    "archivo": nombre_archivo,
                    "status": "success",
                    "uuid": resultado['uuid']
                })
            elif resultado['status'] == 'error' and 'ya existe' in resultado.get('mensaje', ''):
                resultados['duplicados'] += 1
                resultados['detalles'].append({
                    "archivo": nombre_archivo,
                    "status": "error",
                    "mensaje": resultado['mensaje']
                })
            else:
                resultados['errores'] += 1
                resultados['detalles'].append({
                    "archivo": nombre_archivo,
                    "status": "error",
                    "mensaje": resultado.get('mensaje', 'Error desconocido')
                })

        return Response(resultados)

    @action(detail=True, methods=['post'], url_path='generar-poliza')
    def generar_poliza(self, request, pk=None):
        """
        Genera una póliza basada en una plantilla para esta factura.
        """
        from .services.provisioning import generar_poliza_from_factura
        from .models_automation import PlantillaAsiento
        from .serializers import PolizaSerializer
        
        factura = self.get_object()
        plantilla_id = request.data.get('plantilla_id')
        
        if not plantilla_id:
             return Response({"detalle": "Se requiere plantilla_id"}, status=status.HTTP_400_BAD_REQUEST)
             
        try:
            plantilla = PlantillaAsiento.objects.get(pk=plantilla_id)
            poliza = generar_poliza_from_factura(factura, plantilla)
            return Response(PolizaSerializer(poliza).data)
        except Exception as e:
            return Response({"detalle": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='download-diot')
    def download_diot(self, request):
        """
        Genera y descarga el TXT de la DIOT.
        Params: start_date, end_date
        """
        from django.http import HttpResponse
        from .services.diot_service import DIOTService
        from decimal import Decimal
        from types import SimpleNamespace
        
        start = request.query_params.get('start_date')
        end = request.query_params.get('end_date')
        
        # Simulación de obtención de pagos (Legacy Logic preserved)
        # En producción, esto vendría de Egreso.objects.filter(fecha__range=[start, end])
        proveedores = Vendedor.objects.filter(activo=True)
        pagos_para_diot = []
        
        for prov in proveedores:
            # Simulación: Si RFC coincide con dummy, agregamos pago
            if prov.rfc == 'XAXX010101000':
                pago = SimpleNamespace()
                pago.proveedor = prov
                pago.monto = Decimal('1160.00')
                pago.tasa_iva = Decimal('0.16')
                pagos_para_diot.append(pago)
            # Podríamos agregar más lógica simulada o real aquí
        
        txt_content = DIOTService.generar_reporte(pagos_para_diot)
        
        response = HttpResponse(txt_content, content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename="DIOT_Dic2025.txt"'
        return response

    @action(detail=False, methods=['get'], url_path='download-catalogo')
    def download_catalogo(self, request):
        from .services.sat_xml import generate_catalogo_xml
        from django.http import HttpResponse
        
        try:
             anio = int(request.query_params.get('anio', 2023))
             mes = int(request.query_params.get('mes', 1))
             xml_content = generate_catalogo_xml(anio, mes)
             
             response = HttpResponse(xml_content, content_type='text/xml')
             response['Content-Disposition'] = f'attachment; filename="Catalogo_{anio}{mes:02d}.xml"'
             return response
        except Exception as e:
             return Response({"detalle": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='download-balanza')
    def download_balanza(self, request):
        from .services.sat_xml import generate_balanza_xml
        from django.http import HttpResponse
        
        try:
             anio = int(request.query_params.get('anio', 2023))
             mes = int(request.query_params.get('mes', 1))
             
             xml_content = generate_balanza_xml(anio, mes)
             
             response = HttpResponse(xml_content, content_type='text/xml')
             # Format filename per SAT: RFC+Anio+Mes+B+N.xml
             response['Content-Disposition'] = f'attachment; filename="Balanza_{anio}{mes:02d}.xml"'
             return response
        except Exception as e:
             return Response({"detalle": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class PlantillaAsientoViewSet(ContabilidadBaseViewSet):
    from .models_automation import PlantillaAsiento
    from .serializers import PlantillaAsientoSerializer
    queryset = PlantillaAsiento.objects.all().order_by("nombre")
    serializer_class = PlantillaAsientoSerializer

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def strategic_dashboard(request):
    # --- 1. Leer y procesar filtros ---
    timeframe = request.query_params.get("timeframe", "month")
    project_ids_str = request.query_params.get("projects", "all")

    project_ids = []
    if project_ids_str and project_ids_str != "all":
        try:
            project_ids = [int(pid) for pid in project_ids_str.split(",")]
        except (ValueError, TypeError):
            project_ids = []

    # --- 2. NUEVO: Calcular el rango de fechas basado en el timeframe ---
    today = timezone.now().date()
    start_date = today
    end_date = today # Por defecto para 'day', aunque no lo uses.

    if timeframe == "year":
        # Desde el primer día del año actual hasta hoy
        start_date = today.replace(month=1, day=1)
    elif timeframe == "month":
        # Desde el primer día del mes actual hasta hoy
        start_date = today.replace(day=1)
    elif timeframe == "week":
        # Desde el inicio de la semana actual (lunes) hasta hoy
        start_date = today - timedelta(days=today.weekday())

    # --- 3. MODIFICADO: Aplicar filtros a los Querysets base ---
    # Nota: Usamos el manager por defecto que YA FILTRA por is_active=True
    contratos_base = Contrato.objects.filter(
        fecha__range=[start_date, today]
    )
    pagos_base = Pago.objects.filter(
        fecha_pago__range=[start_date, today]
    )

    if project_ids:
        contratos_base = contratos_base.filter(
            presupuesto__upe__proyecto__id__in=project_ids
        )
        pagos_base = pagos_base.filter(
            contrato__presupuesto__upe__proyecto__id__in=project_ids
        )

    # --- 4. Calcular KPIs (ahora se basan en los datos filtrados por fecha) ---
    total_ventas = contratos_base.aggregate(
        total=Coalesce(Sum("monto_mxn"), Decimal("0.0"), output_field=DecimalField())
    )["total"]

    total_recuperado = pagos_base.aggregate(
        total=Coalesce(Sum("monto"), Decimal("0.0"), output_field=DecimalField())
    )["total"]

    # Lógica de KPIs sin cambios, pero ahora usan los querysets filtrados
    total_vencido = Decimal("0.0")
    monto_por_cobrar = total_ventas - total_recuperado
    
    # Manager por defecto filtra activos
    total_upes_query = UPE.objects.all() 
    if project_ids:
        total_upes_query = total_upes_query.filter(proyecto__id__in=project_ids)

    kpis = {
        "upes_total": total_upes_query.count(),
        "ventas": total_ventas,
        "recuperado": total_recuperado,
        "por_cobrar": monto_por_cobrar,
        "vencido": total_vencido,
    }

    # --- 5. Preparar datos para la Gráfica (lógica sin cambios) ---
    if timeframe == "year":
        trunc_func = TruncMonth
        date_format_str = "%b %Y"  # Formato más legible: Ene 2025
    elif timeframe == "week":
        trunc_func = TruncWeek
        date_format_str = "%G-W%V"
    else:  # month (default)
        trunc_func = TruncDay
        date_format_str = "%d-%b" # 03-Oct

    ventas_por_periodo = (
        contratos_base.annotate(periodo=trunc_func("fecha"))
        .values("periodo")
        .annotate(total=Sum("monto_mxn"))
        .order_by("periodo")
    )

    recuperado_por_periodo = (
        pagos_base.annotate(periodo=trunc_func("fecha_pago"))
        .values("periodo")
        .annotate(total=Sum("monto"))
        .order_by("periodo")
    )

    datos_combinados = defaultdict(
        lambda: {"ventas": Decimal("0.0"), "recuperado": Decimal("0.0")}
    )

    for v in ventas_por_periodo:
        if v["periodo"] and v["total"]:
            label = v["periodo"].strftime(date_format_str)
            datos_combinados[label]["ventas"] += v["total"]

    for r in recuperado_por_periodo:
        if r["periodo"] and r["total"]:
            label = r["periodo"].strftime(date_format_str)
            datos_combinados[label]["recuperado"] += r["total"]

    labels_ordenados = sorted(datos_combinados.keys())

    ventas_list = [datos_combinados[label]["ventas"] for label in labels_ordenados]
    recuperado_list = [
        datos_combinados[label]["recuperado"] for label in labels_ordenados
    ]
    programado_list = [
        datos_combinados[label]["ventas"] - datos_combinados[label]["recuperado"]
        for label in labels_ordenados
    ]

    chart_data = {
        "labels": labels_ordenados,
        "ventas": ventas_list,
        "recuperado": recuperado_list,
        "programado": programado_list,
    }

    # --- 6. Ensamblar la respuesta final ---
    data = {
        "kpis": kpis,
        "chart": chart_data,
        "filters": {
            "timeframe": timeframe,
            "projects": project_ids_str,
        },
    }
    return Response(data)


class CertificadoDigitalViewSet(ContabilidadBaseViewSet):
    queryset = CertificadoDigital.objects.all().order_by("-fecha_fin_validez")
    serializer_class = CertificadoDigitalSerializer


class ReporteFinancieroViewSet(viewsets.ViewSet):
    """
    Endpoints para reportes contables financieros.
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def balanza_comprobacion(self, request):
        fecha_start = parse_date(request.query_params.get('start', ''))
        fecha_end = parse_date(request.query_params.get('end', ''))
        
        if not fecha_start or not fecha_end:
            return Response({"error": "Fechas start y end requeridas (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)
            
        data = ReporteFinancieroService.obtener_balanza_comprobacion(fecha_start, fecha_end)
        return Response(data)

    @action(detail=False, methods=['get'])
    def estado_resultados(self, request):
        fecha_start = parse_date(request.query_params.get('start', ''))
        fecha_end = parse_date(request.query_params.get('end', ''))
        
        if not fecha_start or not fecha_end:
            return Response({"error": "Fechas start y end requeridas (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        data = ReporteFinancieroService.obtener_estado_resultados(fecha_start, fecha_end)
        return Response(data)

    @action(detail=False, methods=['get'])
    def balance_general(self, request):
        fecha_corte = parse_date(request.query_params.get('date', ''))
        
        if not fecha_corte:
            return Response({"error": "Fecha date requerida (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        data = ReporteFinancieroService.obtener_balance_general(fecha_corte)
        return Response(data)


class BuzonMensajeViewSet(ContabilidadBaseViewSet):
    """
    Vista del Buzón Tributario.
    """
    queryset = BuzonMensaje.objects.all().order_by("-fecha_recibido")
    serializer_class = BuzonMensajeSerializer
    
    @action(detail=False, methods=['post'], url_path='sincronizar')
    def sincronizar(self, request):
        """
        Fuerza la sincronización con el SAT para buscar nuevos mensajes.
        """
        rfc = request.data.get('rfc')
        if not rfc:
            return Response({"error": "RFC es requerido"}, status=status.HTTP_400_BAD_REQUEST)
            
        mensajes = SATIntegrationService.sincronizar_buzon(rfc)
        serializer = self.get_serializer(mensajes, many=True)
        return Response(serializer.data)

class OpinionCumplimientoViewSet(ContabilidadBaseViewSet):
    """
    Vista de Opinión de Cumplimiento.
    """
    queryset = OpinionCumplimiento.objects.all().order_by("-fecha_consulta")
    serializer_class = OpinionCumplimientoSerializer
    
    @action(detail=False, methods=['post'], url_path='consultar')
    def consultar(self, request):
        """
        Consulta en tiempo real.
        """
        rfc = request.data.get('rfc')
        if not rfc:
             return Response({"error": "RFC es requerido"}, status=status.HTTP_400_BAD_REQUEST)
             
        opinion = SATIntegrationService.consultar_opinion_cumplimiento(rfc)
        return Response(self.get_serializer(opinion).data)
